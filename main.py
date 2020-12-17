"""
Author: CHR-onicles©
Date: 05/12/2020

DSC UG Python Challenge - Time Tracker App.
"""

from tkinter import *
from tkcalendar import *
from tkinter.filedialog import askdirectory
from tkinter import messagebox
from PIL import Image, ImageTk
from datetime import datetime as dt
import openpyxl
import os
import threading



# MAIN WINDOW CONFIGURATIONS-------------------------------------------------------------------

root = Tk()
root.title('TimeTracker')
root.iconbitmap('images/cat.ico')

SCREEN_HEIGHT = root.winfo_screenheight()
SCREEN_WIDTH = root.winfo_screenwidth()

# print(SCREEN_HEIGHT, SCREEN_WIDTH)  # For debugging
ROOT_WIDTH = 550
ROOT_HEIGHT = 640

# x and y coordinates to center app on screen
x_pos = (SCREEN_WIDTH/2) - (ROOT_WIDTH/2)
y_pos = (SCREEN_HEIGHT/2) - (ROOT_HEIGHT/2)

root.geometry(f'{ROOT_WIDTH}x{ROOT_HEIGHT}+{int(x_pos)}+{int(y_pos)}')

# Preventing user from increasing or reducing app size as most widgets are statically placed!
root.resizable(False, False)
root.configure(bg='#5fc29e')  # turquoise-ish colour used in status bar and calendar.



# GLOBAL VARIABLES-----------------------------------------------------------------------

# Widget Position Coordinates:
# Duplicates allowed in case there is a need to change specific widgets coords in future.

# Top Canvas Stuff
t_text_x, t_text_y = 270, 30
t_date_text_x, t_date_text_y = 70, 100
t_time_text_x, t_time_text_y = 225, 100
t_cal_button_x, t_cal_button_y = 136, 100
t_cf_button_x, t_cf_button_y = 490, 90
t_cx_button_x, t_cx_button_y = 490, 160
t_time_now_button_x, t_time_now_button_y = 415, 100

# Middle Canvas Stuff
m_text_x, m_text_y = 280, 26
m_date_text_x, m_date_text_y = 70, 100
m_time_text_x, m_time_text_y = 225, 100
m_cal_button_x, m_cal_button_y = 136, 100
m_cf_button_x, m_cf_button_y = 490, 90
m_cx_button_x, m_cx_button_y = 490, 160
m_time_now_button_x, m_time_now_button_y = 415, 100

# Bottom Canvas Stuff
bottom_text_x, bottom_text_y = 280, 30
cbutton_x, cbutton_y = 275, 100
rbutton_x, rbutton_y = 60, 160
sbutton_x, sbutton_y = 490, 160
amt_lbl_x, amt_lbl_y = 270, 150


# Calendar stuff
first_time_open = False
date_from_calendar = ''
date_from_calendar2 = ''

# Status bar text
rate_text = 'Rate: 1hr = $5.00'

# Payment calculation variable
amount_to_be_paid = 0.0
total_hours_spent = 0.0

# Excel stuff variable
excel_first_time_open = 0
alpha_current_dir = os.getcwd()
dir_location = ''
# -------------------------------------------------------------------------------------


# DEFINING BINDING EVENT FUNCTIONS

def spin_box_hover_in(event, spin_box):
    """
    Function to change the spinbox's foreground colour and status bar when hovered on.

    :param event: necessary argument in order for binding event to work.
    :param spin_box: spin box object that is passed in to be updated.
    """
    spin_box.config(fg='black')

    if spin_box == top_hour_spin_box or spin_box == m_hour_spin_box:
        status_bar_left.config(text='Set the hour...', font=('consolas', 12, 'italic'))

    elif spin_box == top_minute_spin_box or spin_box == m_minute_spin_box:
        status_bar_left.config(text='Set the minutes...', font=('consolas', 12, 'italic'))


def spin_box_hover_out(event, spin_box):
    """
    Function to change the spinbox's foreground color back to normal.

    :param event: necessary argument in order for binding event to work.
    :param spin_box: spin box object that is passed in to be updated.
    """
    spin_box.config(fg='gray')
    status_bar_left.config(text=rate_text, font=('consolas', 12))


def submit_button_hover_in(event):
    """
    Binding event function to update submit button in calendar,
    when it is hovered on.
    """
    submit_button.config(bg='white', font=('consolas', 14, 'italic'))


def submit_button_hover_out(event):
    """
    Binding event function to update submit button in calendar,
    when it is hovered out of.
    """
    submit_button.config(bg='SystemButtonFace', font=('consolas', 14))


def calculate_payment_button_hover_in(event):
    """
    Function to change colour and font style of calculate payment button and status bar
    when hovered on.
    """
    calc_payment_button.config(font=('consolas', 22))
    status_bar_left.config(text='Calculate payment based on info provided...', font=('consolas', 10, 'italic'))


def calculate_payment_button_hover_out(event):
    """
    Function to change calculate payment button's colour, font style and status bar
    back to normal when mouse pointer is no longer hovering on the button.
    """
    calc_payment_button.config(font=('consolas', 20), bg='SystemButtonFace')
    status_bar_left.config(text=rate_text, font=('consolas', 12))


def save_button_hover_in(event):
    """
    Binding event function to update save button's color, size and status bar
    when mouse pointer hovers on it.
    """
    global save_icon, save_button, sb_window
    bottom_canvas.delete(sb_window)
    save_icon = ImageTk.PhotoImage(Image.open('images/save.png').resize((70, 65), Image.ANTIALIAS))
    save_button = Button(bottom_canvas, image=save_icon, command=save_to_excel)
    sb_window = bottom_canvas.create_window(sbutton_x, sbutton_y, window=save_button)
    save_button.bind('<Leave>', save_button_hover_out)

    # Update status bar
    status_bar_left.config(text='Save payment info to excel file...', font=('consolas', 11, 'italic'))


def save_button_hover_out(event):
    """
    Binding event function to return save button to normal when hovered out of.
    """
    global save_icon, save_button, sb_window
    bottom_canvas.delete(sb_window)
    save_icon = ImageTk.PhotoImage(Image.open('images/save.png').resize((65, 60), Image.ANTIALIAS))
    save_button = Button(bottom_canvas, image=save_icon, command=save_to_excel)
    sb_window = bottom_canvas.create_window(sbutton_x, sbutton_y, window=save_button)
    save_button.bind('<Enter>', save_button_hover_in)

    # Update status bar again
    status_bar_left.config(text=rate_text, font=('consolas', 12))


def reset_button_hover_in(event):
    """
    Function to update reset button and status bar when hovered on.
    """
    global reset_icon, reset_button, r_window
    bottom_canvas.delete(r_window)
    reset_icon = ImageTk.PhotoImage(Image.open('images/Reset.png').resize((65, 65), Image.ANTIALIAS))
    reset_button = Button(bottom_canvas, image=reset_icon, command=reset_values)
    r_window = bottom_canvas.create_window(rbutton_x, rbutton_y, window=reset_button)
    reset_button.bind('<Leave>', reset_button_hover_out)

    # Update status bar
    status_bar_left.config(text='Reset calculation...', font=('consolas', 12, 'italic'))


def reset_button_hover_out(event):
    """
    Function to update reset button and status bar back to normal when hovered out of.
    """
    global reset_icon, reset_button, r_window
    bottom_canvas.delete(r_window)
    reset_icon = ImageTk.PhotoImage(Image.open('images/Reset.png').resize((60, 60), Image.ANTIALIAS))
    reset_button = Button(bottom_canvas, image=reset_icon, command=reset_values)
    r_window = bottom_canvas.create_window(rbutton_x, rbutton_y, window=reset_button)
    reset_button.bind('<Enter>', reset_button_hover_in)

    # Update status bar
    status_bar_left.config(text=rate_text, font=('consolas', 12))


def calendar_button_hover_in(event, canvas):
    """
    Function to update the calendar button when it is hovered on.
    """

    # Created 2 separate instances of the hover event because of conflict between them when they both use the same
    # resources.
    if canvas == top_canvas:  # this is top canvas
        global cl_window, calendar_button, cal_icon
        canvas.delete(cl_window)
        cal_icon = ImageTk.PhotoImage(Image.open('images/cal2_icon.png').resize((50, 50), Image.ANTIALIAS))
        calendar_button = Button(canvas, image=cal_icon, bg='light gray', command=lambda: open_calendar(canvas))
        cl_window = canvas.create_window(t_cal_button_x, t_cal_button_y, window=calendar_button)
        calendar_button.bind('<Leave>', lambda e: calendar_button_hover_out(e, canvas))

    else:
        global cl_window2, cal_icon2, calendar_button2
        canvas.delete(cl_window2)
        cal_icon2 = ImageTk.PhotoImage(Image.open('images/cal2_icon.png').resize((50, 50), Image.ANTIALIAS))
        calendar_button2 = Button(canvas, image=cal_icon2, bg='light gray', command=lambda: open_calendar(canvas))
        cl_window2 = canvas.create_window(m_cal_button_x, m_cal_button_y, window=calendar_button2)
        calendar_button2.bind('<Leave>', lambda e: calendar_button_hover_out(e, canvas))

    status_bar_left.config(text='Pick a date from the calendar...', font=('consolas', 12, 'italic'))


def calendar_button_hover_out(event, canvas):
    """
    Function to update the calendar button when it is hovered out of.
    """
    if canvas == top_canvas:  # top canvas
        global cl_window, calendar_button, cal_icon
        canvas.delete(cl_window)
        cal_icon = ImageTk.PhotoImage(Image.open('images/cal2_icon.png').resize((45, 45), Image.ANTIALIAS))
        calendar_button = Button(canvas, image=cal_icon, bg='light gray', command=lambda: open_calendar(canvas))
        cl_window = canvas.create_window(t_cal_button_x, t_cal_button_y, window=calendar_button)
        calendar_button.bind('<Enter>', lambda e: calendar_button_hover_in(e, canvas))

    else:  # middle canvas
        global cl_window2, cal_icon2, calendar_button2
        canvas.delete(cl_window2)
        cal_icon2 = ImageTk.PhotoImage(Image.open('images/cal2_icon.png').resize((45, 45), Image.ANTIALIAS))
        calendar_button2 = Button(canvas, image=cal_icon2, bg='light gray', command=lambda: open_calendar(canvas))
        cl_window2 = canvas.create_window(m_cal_button_x, m_cal_button_y, window=calendar_button2)
        calendar_button2.bind('<Enter>', lambda e: calendar_button_hover_in(e, canvas))

    status_bar_left.config(text=rate_text, font=('consolas', 12))


def confirm_button_hover_in(event, canvas):
    """
    Function to update the confirm button when it is hovered on.
    """
    if canvas == top_canvas:
        global c_icon, check_button, cf_window
        canvas.delete(cf_window)
        c_icon = ImageTk.PhotoImage(Image.open('images/check_icon.png').resize((65, 65), Image.ANTIALIAS))
        check_button = Button(canvas, image=c_icon, command=lambda: confirm_button_click(canvas))
        cf_window = canvas.create_window(t_cf_button_x, t_cf_button_y, window=check_button)
        check_button.bind('<Leave>', lambda e: confirm_button_hover_out(e, canvas))

    else:  # middle canvas
        global c_icon2, check_button2, cf_window2
        canvas.delete(cf_window2)
        c_icon2 = ImageTk.PhotoImage(Image.open('images/check_icon.png').resize((65, 65), Image.ANTIALIAS))
        check_button2 = Button(canvas, image=c_icon2, command=lambda: confirm_button_click(canvas))
        cf_window2 = canvas.create_window(m_cf_button_x, m_cf_button_y, window=check_button2)
        check_button2.bind('<Leave>', lambda e: confirm_button_hover_out(e, canvas))

    status_bar_left.config(text='Confirm date and time selection...', font=('consolas', 11, 'italic'))


def confirm_button_hover_out(event, canvas):
    """
    Function to update the confirm button when it is hovered out of.
    """
    if canvas == top_canvas:  # top canvas
        global c_icon, check_button, cf_window
        canvas.delete(cf_window)
        c_icon = ImageTk.PhotoImage(Image.open('images/check_icon.png').resize((60, 60), Image.ANTIALIAS))
        check_button = Button(canvas, image=c_icon, command=lambda: confirm_button_click(canvas))
        cf_window = canvas.create_window(t_cf_button_x, t_cf_button_y, window=check_button)
        check_button.bind('<Enter>', lambda e: confirm_button_hover_in(e, canvas))

    else:  # middle canvas
        global c_icon2, check_button2, cf_window2
        canvas.delete(cf_window2)
        c_icon2 = ImageTk.PhotoImage(Image.open('images/check_icon.png').resize((60, 60), Image.ANTIALIAS))
        check_button2 = Button(canvas, image=c_icon2, command=lambda: confirm_button_click(canvas))
        cf_window2 = canvas.create_window(m_cf_button_x, m_cf_button_y, window=check_button2)
        check_button2.bind('<Enter>', lambda e: confirm_button_hover_in(e, canvas))

    status_bar_left.config(text=rate_text, font=('consolas', 12))


def cancel_button_hover_in(event, canvas):
    """
    Function to update the cancel button when it is hovered on.
    """
    if canvas == top_canvas:
        global cancel_icon, cancel_button, cx_window
        canvas.delete(cx_window)
        cancel_icon = ImageTk.PhotoImage(Image.open('images/remove_icon.jpg').resize((65, 65), Image.ANTIALIAS))
        cancel_button = Button(canvas, image=cancel_icon, command=lambda: cancel_button_click(canvas))
        cx_window = canvas.create_window(t_cx_button_x, t_cx_button_y, window=cancel_button)
        cancel_button.bind('<Leave>', lambda e: cancel_button_hover_out(e, canvas))

    else:  # middle canvas
        global cancel_icon2, cancel_button2, cx_window2
        canvas.delete(cx_window2)
        cancel_icon2 = ImageTk.PhotoImage(Image.open('images/remove_icon.jpg').resize((65, 65), Image.ANTIALIAS))
        cancel_button2 = Button(canvas, image=cancel_icon2, command=lambda: cancel_button_click(canvas))
        cx_window2 = canvas.create_window(m_cx_button_x, m_cx_button_y, window=cancel_button2)
        cancel_button2.bind('<Leave>', lambda e: cancel_button_hover_out(e, canvas))

    status_bar_left.config(text='Cancel date and time selected...', font=('consolas', 11, 'italic'))


def cancel_button_hover_out(event, canvas):
    """
    Function to update the cancel button when it is hovered out of.
    """
    if canvas == top_canvas:
        global cancel_icon, cancel_button, cx_window
        canvas.delete(cx_window)
        cancel_icon = ImageTk.PhotoImage(Image.open('images/remove_icon.jpg').resize((60, 60), Image.ANTIALIAS))
        cancel_button = Button(canvas, image=cancel_icon, command=lambda: cancel_button_click(canvas))
        cx_window = canvas.create_window(t_cx_button_x, t_cx_button_y, window=cancel_button)
        cancel_button.bind('<Enter>', lambda e: cancel_button_hover_in(e, canvas))

    else:  # middle canvas
        global cancel_icon2, cancel_button2, cx_window2
        canvas.delete(cx_window2)
        cancel_icon2 = ImageTk.PhotoImage(Image.open('images/remove_icon.jpg').resize((60, 60), Image.ANTIALIAS))
        cancel_button2 = Button(canvas, image=cancel_icon2, command=lambda: cancel_button_click(canvas))
        cx_window2 = canvas.create_window(m_cx_button_x, m_cx_button_y, window=cancel_button2)
        cancel_button2.bind('<Enter>', lambda e: cancel_button_hover_in(e, canvas))

    status_bar_left.config(text=rate_text, font=('consolas', 12))


def time_now_button_hover_in(event, canvas):
    """
    Function to update time_now button when it is hovered on.
    """
    if canvas == top_canvas:
        global time_now_icon, time_now_button, tn_window
        canvas.delete(tn_window)
        time_now_icon = ImageTk.PhotoImage(Image.open('images/t_now.png').resize((50, 50), Image.ANTIALIAS))
        time_now_button = Button(canvas, image=time_now_icon, command=lambda: get_current_date_time(canvas))
        tn_window = canvas.create_window(t_time_now_button_x, t_time_now_button_y, window=time_now_button)
        time_now_button.bind('<Leave>', lambda e: time_now_button_hover_out(e, canvas))

    else:  # middle canvas
        global time_now_icon2, time_now_button2, tn_window2
        canvas.delete(tn_window2)
        time_now_icon2 = ImageTk.PhotoImage(Image.open('images/t_now.png').resize((50, 50), Image.ANTIALIAS))
        time_now_button2 = Button(canvas, image=time_now_icon2,
                                  command=lambda: get_current_date_time(canvas))
        tn_window2 = canvas.create_window(m_time_now_button_x, m_time_now_button_y, window=time_now_button2)
        time_now_button2.bind('<Leave>', lambda e: time_now_button_hover_out(e, canvas))

    status_bar_left.config(text='Get current date and time...', font=('consolas', 12, 'italic'))


def time_now_button_hover_out(event, canvas):
    """
    Function to update time now button when it is hovered out of.
    """
    if canvas == top_canvas:
        global time_now_icon, time_now_button, tn_window
        canvas.delete(tn_window)
        time_now_icon = ImageTk.PhotoImage(Image.open('images/t_now.png').resize((45, 45), Image.ANTIALIAS))
        time_now_button = Button(canvas, image=time_now_icon, command=lambda: get_current_date_time(canvas))
        tn_window = canvas.create_window(t_time_now_button_x, t_time_now_button_y, window=time_now_button)
        time_now_button.bind('<Enter>', lambda e: time_now_button_hover_in(e, canvas))

    else:  # middle canvas
        global time_now_icon2, time_now_button2, tn_window2
        canvas.delete(tn_window2)
        time_now_icon2 = ImageTk.PhotoImage(Image.open('images/t_now.png').resize((45, 45), Image.ANTIALIAS))
        time_now_button2 = Button(canvas, image=time_now_icon2,
                                  command=lambda: get_current_date_time(canvas))
        tn_window2 = canvas.create_window(m_time_now_button_x, m_time_now_button_y, window=time_now_button2)
        time_now_button2.bind('<Enter>', lambda e: time_now_button_hover_in(e, canvas))

    status_bar_left.config(text=rate_text, font=('consolas', 12))



# DEFINING COMMAND FUNCTIONS--------------------------------------------------------

def open_calendar(canvas):
    """
    Command function to open calendar for user to choose a date.

    :param: canvas: canvas object for the function to know which canvas called it and
                    give it specific attributes.
    """

    global first_time_open, cal_window

    def create_calendar_window():
        global first_time_open, cal_window

        # Grabbing current date from system
        year = dt.now().year
        month = dt.now().month
        day = dt.now().day

        # Calendar window configurations
        cal_window = Toplevel()
        cal_window.title('Calendar')
        cal_window.iconbitmap('images/cat.ico')

        # To keep the calendar window within the app screen (specifically top left of app)
        cal_window.geometry(f'250x270+{int(x_pos)}+{int(y_pos)}')
        cal_window.resizable(False, False)
        cal_window.configure(bg='#5fc29e')

        def get_date():
            """
            Based on the canvas argument from outer function, this function
            updates a global variable on the date got from the calendar.
            """
            global cal_window, selected_label, date_from_calendar, date_from_calendar2
            if canvas == top_canvas:
                date_from_calendar = cal.get_date()

                # rearranging of date format to: DD/MM/YY
                temp = date_from_calendar.split('/')

                # Prepending 0 to single digits
                if len(temp[0]) == 1 and len(temp[1]) == 1:
                    temp[0] = '0' + temp[0]
                    temp[1] = '0' + temp[1]
                elif len(temp[0]) == 1:
                    temp[0] = '0' + temp[0]
                elif len(temp[1]) == 1:
                    temp[1] = '0' + temp[1]

                temp2 = temp[1] + '/' + temp[0] + '/' + temp[2]
                date_from_calendar = temp2

                # Close calendar window 2 seconds after user clicks submit
                cal_window.after(2000, cal_window.destroy)

            elif canvas == middle_canvas:
                date_from_calendar2 = cal.get_date()

                # rearranging of date format to: DD/MM/YY
                temp = date_from_calendar2.split('/')

                # Prepending 0 to single digits
                if len(temp[0]) == 1 and len(temp[1]) == 1:
                    temp[0] = '0' + temp[0]
                    temp[1] = '0' + temp[1]
                elif len(temp[0]) == 1:
                    temp[0] = '0' + temp[0]
                elif len(temp[1]) == 1:
                    temp[1] = '0' + temp[1]

                temp2 = temp[1] + '/' + temp[0] + '/' + temp[2]
                date_from_calendar2 = temp2

                # Close calendar window 2 seconds after user clicks submit
                cal_window.after(2000, cal_window.destroy)

            selected_label = Label(cal_window, text='Date Selected!', font=('consolas', 14, 'italic'),
                                   bg='#5fc29e')
            selected_label.pack(pady=5)
            submit_button.config(state=DISABLED)

        global cal
        cal = Calendar(cal_window, selectmode='day', year=year, month=month, day=day)
        cal.pack(fill=BOTH, expand=1)

        # Submit button for calendar
        global submit_button
        submit_button = Button(cal_window, text='Submit', font=('consolas', 14), command=get_date)
        submit_button.pack(pady=(5, 0))
        submit_button.bind('<Enter>', submit_button_hover_in)
        submit_button.bind('<Leave>', submit_button_hover_out)
        first_time_open = True

    if first_time_open is False:  # Opened calendar window for the first time
        create_calendar_window()

    else:
        if cal_window.winfo_exists():  # Calendar window is still open
            first_time_open = True
        else:  # Calendar window was opened but closed.
            create_calendar_window()


def status_bar_time_update():
    """
    Function to update the time in the status bar through a thread.
    """
    cur_time = dt.now().strftime('%H:%M:%S')
    status_bar_right.config(text='|System Time(24h): ' + cur_time)
    status_bar_right.after(1000, status_bar_time_update)


def confirm_button_click(canvas):
    """
    Command function to display user's choice of date and time on the canvas provided.

    :param canvas: canvas object to know which canvas called it and attribute specific
                    configurations to it or some other global variables.
    """
    global date_window, date_window2, selected_date_label, s_time, selected_date_label2, s_time2

    def remove_user_prepended_zeros(var):
        """
        Small Helper function to remove user prepended zeros to time because app will take there of that.

        :param var: spin box value that is passed to function.
        :return: returns new var without the prepended zero.
        """
        if len(var.get()) == 2 and var.get()[0] == '0':
            var.set(var.get()[1])
            return var.get()
        else:
            return var.get()  # do nothing

    # Putting all spinbox values in local variables because
    # its easier to perform operations on those variables.
    top_hour = StringVar()
    top_hour.set(top_hour_spin_box.get())
    top_minute = StringVar()
    top_minute.set(top_minute_spin_box.get())
    m_hour = StringVar()
    m_hour.set(m_hour_spin_box.get())
    m_minute = StringVar()
    m_minute.set(m_minute_spin_box.get())



    # Check for user prepended zeros and remove them with function
    top_hour.set(remove_user_prepended_zeros(top_hour))
    top_minute.set(remove_user_prepended_zeros(top_minute))
    m_hour.set(remove_user_prepended_zeros(m_hour))
    m_minute.set(remove_user_prepended_zeros(m_minute))


    if canvas == top_canvas:

        # If user types in more than 2 numbers, generate error message
        if len(top_hour.get()) > 2 or len(top_minute.get()) > 2:
            messagebox.showerror(title='INVALID INPUT', message='TIME ENTERED IS INVALID')

        else:
            # Prepend 0 if its a single digit
            # Decided to use this and not len() because this helps filter out non-digit input from user :)
            if (0 <= int(top_hour.get()) <= 9) and (0 <= int(top_minute.get()) <= 9):
                s_time = '0' + top_hour.get() + ':' + '0' + top_minute.get()
            elif 0 <= int(top_minute.get()) <= 9:
                s_time = top_hour.get() + ':' + '0' + top_minute.get()
            elif 0 <= int(top_hour.get()) <= 9:
                s_time = '0' + top_hour.get() + ':' + top_minute.get()
            else:
                s_time = top_hour.get() + ':' + top_minute.get()
            selected_date_label.config(text='You selected: ' + date_from_calendar + ' ' + s_time)
            date_window = canvas.create_window(225, 170, window=selected_date_label)

    elif canvas == middle_canvas:

        # If user types in more than 2 numbers, generate error message
        if len(m_hour.get()) > 2 or len(m_minute.get()) > 2:
            messagebox.showerror(title='INVALID INPUT', message='TIME ENTERED IS INVALID')

        else:
            if (0 <= int(m_minute.get()) <= 9) and (0 <= int(m_hour.get()) <= 9):
                s_time2 = '0' + m_hour.get() + ':' + '0' + m_minute.get()
            elif 0 <= int(m_minute.get()) <= 9:
                s_time2 = m_hour.get() + ':' + '0' + m_minute.get()
            elif 0 <= int(m_hour.get()) <= 9:
                s_time2 = '0' + m_hour.get() + ':' + m_minute.get()
            else:
                s_time2 = m_hour.get() + ':' + m_minute.get()
            selected_date_label2.config(text='You selected: ' + date_from_calendar2 + ' ' + s_time2)
            date_window2 = canvas.create_window(225, 170, window=selected_date_label2)


def cancel_button_click(canvas):
    """
    Command function to remove label/text telling user his choice of date and time.

    :param canvas: canvas object to know which canvas called it and update specific labels.
    """
    if canvas == middle_canvas:
        canvas.delete(date_window2)

    else:  # top canvas
        canvas.delete(date_window)


def get_current_date_time(canvas):
    """
    Command function to get the system's current date and time.

    :param canvas: canvas object to know which canvas called it and update its labels.
    """
    global s_time, selected_date_label, date_window, date_window2, s_time2
    global date_from_calendar, date_from_calendar2, selected_date_label2

    if canvas == top_canvas:
        date_from_calendar = dt.now().strftime('%d/%m/%y')
        s_time = dt.now().strftime('%H:%M')
        selected_date_label.config(text='You selected: ' + date_from_calendar + ' ' + s_time)
        date_window = canvas.create_window(225, 170, window=selected_date_label)

    else:  # middle canvas
        date_from_calendar2 = dt.now().strftime('%d/%m/%y')
        s_time2 = dt.now().strftime('%H:%M')
        selected_date_label2.config(text='You selected: ' + date_from_calendar2 + ' ' + s_time2)
        date_window2 = canvas.create_window(225, 170, window=selected_date_label2)


def payment_calculation():
    """
    Function to calculate payment amount based on the date and time passed in
    by user.
    Rate: $5 for 60mins (1hr).
    """
    # Do nothing with dates for now since both dates(date started and date completed) will be the same.
    # NB: If User selects different dates, app only checks the time regardless.
    # Hopefully no one enters different dates and get some crazy output :)
    # Hmmm...naa... TODO: make sure User enters same date twice or only display one calendar button for QoL

    global amount_label, calc_payment_button, amount_to_be_paid, s_time, s_time2, total_hours_spent
    amount_label = bottom_canvas.create_text(250, 150, text='')

    # Perform operations on hours and minutes only
    # Grabbing hours and minutes
    hour_completed = int(s_time2[0:2])
    minutes_completed = int(s_time2[3:])
    hour_started = int(s_time[0:2])
    minutes_started = int(s_time[3:])

    hours_elapsed = hour_completed - hour_started
    minutes_elapsed = minutes_completed - minutes_started

    # Convert everything to minutes and perform calculations based on that
    hours_to_mins = hours_elapsed * 60
    total_minutes_elapsed = minutes_elapsed + hours_to_mins
    total_hours_spent = round(total_minutes_elapsed / 60, 2)
    amount_to_be_paid = round(float(total_minutes_elapsed * 1 / 12), 2)

    # hmm...Might generate false positive errors since user can choose different dates
    # TODO: Do the TODO just before this one
    if total_minutes_elapsed < 0:
        messagebox.showerror(title='INVALID ARGUMENT', message='INVALID INFORMATION ENTERED!')
        calc_payment_button.config(state=DISABLED)
    else:
        amount_label = bottom_canvas.create_text(amt_lbl_x, amt_lbl_y, text='$' + str(amount_to_be_paid),
                                                 font=('consolas', 40, 'bold'), fill='#e8ebea')
        calc_payment_button.config(state=DISABLED)


def reset_values():
    """
    Function to reset dates, time and amount calculated.
    """
    global calc_payment_button
    bottom_canvas.delete(amount_label)
    middle_canvas.delete(date_window2)
    top_canvas.delete(date_window)
    calc_payment_button.config(state=NORMAL)


def save_to_excel():
    """
    Function to save date/time started and completed and payment calculated
    to an excel file.
    """
    global excel_first_time_open, amount_to_be_paid, dir_location, total_hours_spent
    program_dir = alpha_current_dir


    def check_total_hours_spent():
        """
        Small Helper function to check if the total hours spent is
         and bring up an error box.
        """
        if total_hours_spent < 0:
            messagebox.showerror(title='INVALID ARGUMENT', message='INVALID INFORMATION ENTERED!')
            return False
        else:
            return True


    def create_new_file():
        """
        Function to create new excel file and input information from app.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook['Sheet']

        # Column headers
        sheet['A1'] = 'Date Started'
        sheet['B1'] = 'Time Started'
        sheet['C1'] = 'Date Completed'
        sheet['D1'] = 'Time Completed'
        sheet['E1'] = 'Total Hours Spent'
        sheet['F1'] = 'Calculated Payment'
        sheet['G1'] = 'Time Info was Saved'

        # Increase width for better view
        for cols in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            sheet.column_dimensions[cols].width = 20

        # This is the first time so we can write specifically to certain cells
        sheet['A2'] = date_from_calendar
        sheet['B2'] = s_time
        sheet['C2'] = date_from_calendar2
        sheet['D2'] = s_time2
        sheet['E2'] = total_hours_spent
        sheet['F2'] = '$' + str(amount_to_be_paid)
        sheet['G2'] = dt.now().strftime('%d/%m/%Y, %H:%M:%S')

        # Save excel file and close it
        workbook.save('payment_history.xlsx')
        workbook.close()

        # Change back to program's directory
        os.chdir(program_dir)

        # Message box for successful entry to NEW file
        messagebox.showinfo(title='Save To Excel File', message='Information saved successfully to NEW excel file!')

    def edit_existing_file():
        """
        Function to edit already existing excel file and input info from app.
        """
        workbook = openpyxl.load_workbook('payment_history.xlsx')
        sheet = workbook['Sheet']

        found_available_cell = False
        row_num = 0
        for row in sheet.iter_rows(min_row=1, max_row=200, min_col=1, max_col=6):
            row_num += 1
            for count, cell in enumerate(row):

                # if all cells in a row are empty, then that row is available for writing
                if cell.value is None and cell == row[-1]:
                    sheet.cell(row=row_num, column=1).value = date_from_calendar
                    sheet.cell(row=row_num, column=2).value = s_time
                    sheet.cell(row=row_num, column=3).value = date_from_calendar2
                    sheet.cell(row=row_num, column=4).value = s_time2
                    sheet.cell(row=row_num, column=5).value = total_hours_spent
                    sheet.cell(row=row_num, column=6).value = '$' + str(amount_to_be_paid)
                    sheet.cell(row=row_num, column=7).value = dt.now().strftime('%d/%m/%Y, %H:%M:%S')
                    found_available_cell = True
                    break
                elif cell.value is not None:  # if any cell contains information, break
                    break
            if found_available_cell is True:
                break

        # Save and close excel file
        workbook.save('payment_history.xlsx')
        workbook.close()

        # Change back to program's directory
        os.chdir(program_dir)

        # Message box for successful entry to EXISTING file
        messagebox.showinfo(title='Save To Excel File', message='Information saved successfully to EXISTING excel file!')


    # Main Excel Logic
    if excel_first_time_open == 0:  # Save button has not been clicked already
        if check_total_hours_spent() is True:
            dir_location = askdirectory(initialdir='.\\', title='Select Directory To Save To')

            # Change location to the user-chosen one
            os.chdir(dir_location)

            # Search for already existing file
            for item in os.listdir():
                if item != 'payment_history.xlsx' and item == os.listdir()[-1]:
                    create_new_file()

                elif item == 'payment_history.xlsx':
                    edit_existing_file()
                    break
        else:
            pass  # Show error box and do nothing
    else:
        if check_total_hours_spent() is True:
            # Change location to the user-chosen one
            os.chdir(dir_location)
            edit_existing_file()
        else:
            pass  # Show error box and do nothing

    excel_first_time_open += 1




# TOP CANVAS---------------------------------------------------------------------------------------

top_canvas = Canvas(root, width=ROOT_WIDTH, height=200, borderwidth=0)
top_canvas.grid(row=0, column=0, sticky=W + E, columnspan=2)
img1 = ImageTk.PhotoImage(Image.open('images/clock_dark.jfif'), Image.ANTIALIAS)
top_canvas.create_image(0, 0, image=img1, anchor=NW)
top_canvas.create_text(t_text_x, t_text_y, text='Date/Time Started', font=('android 7', 25))

# Top Canvas Hour Spin Box
top_hour_spin_box = Spinbox(top_canvas, from_=0, to=23, width=2, font=('consolas', 20), fg='gray')
top_canvas.create_window(290, 100, window=top_hour_spin_box)
top_hour_spin_box.bind('<Enter>', lambda event: spin_box_hover_in(event, top_hour_spin_box))
top_hour_spin_box.bind('<Leave>', lambda event: spin_box_hover_out(event, top_hour_spin_box))

# Colon separating minute from hours
top_canvas.create_text(323, 100, text=':', font=('android 7', 25, 'bold'), fill='white')

# Top Canvas Minute Spin Box
top_minute_spin_box = Spinbox(top_canvas, from_=0, to=59, width=2, font=('consolas', 20), fg='gray')
top_canvas.create_window(355, 100, window=top_minute_spin_box)
top_minute_spin_box.bind('<Enter>', lambda event: spin_box_hover_in(event, top_minute_spin_box))
top_minute_spin_box.bind('<Leave>', lambda event: spin_box_hover_out(event, top_minute_spin_box))

# Time text
top_canvas.create_text(t_time_text_x, t_time_text_y, text='Time:', font=('consolas', 20))

# Date text
top_canvas.create_text(t_date_text_x, t_date_text_y, text='Date:', font=('consolas', 20))

# Calendar button for top canvas
t_cal_icon = ImageTk.PhotoImage(Image.open('images/cal2_icon.png').resize((45, 45), Image.ANTIALIAS))
t_calendar_button = Button(top_canvas, image=t_cal_icon, bg='light gray', command=lambda: open_calendar(top_canvas))
cl_window = top_canvas.create_window(t_cal_button_x, t_cal_button_y, window=t_calendar_button)
t_calendar_button.bind('<Enter>', lambda event: calendar_button_hover_in(event, top_canvas))
t_calendar_button.bind('<Leave>', lambda event: calendar_button_hover_out(event, top_canvas))

# Selected date and time text
s_time = ''
selected_date_label = Label(top_canvas, text='You selected: ' + date_from_calendar + ' ' + s_time,
                            font=('consolas', 20, 'italic', 'bold'))

# Checkmark button
c_icon = ImageTk.PhotoImage(Image.open('images/check_icon.png').resize((60, 60), Image.ANTIALIAS))
check_button = Button(top_canvas, image=c_icon, command=lambda: confirm_button_click(top_canvas))
cf_window = top_canvas.create_window(t_cf_button_x, t_cf_button_y, window=check_button)
check_button.bind('<Enter>', lambda event: confirm_button_hover_in(event, top_canvas))
check_button.bind('<Leave>', lambda event: confirm_button_hover_out(event, top_canvas))

# Cancel button
cancel_icon = ImageTk.PhotoImage(Image.open('images/remove_icon.jpg').resize((60, 60), Image.ANTIALIAS))
cancel_button = Button(top_canvas, image=cancel_icon, command=lambda: cancel_button_click(top_canvas))
cx_window = top_canvas.create_window(t_cx_button_x, t_cx_button_y, window=cancel_button)
cancel_button.bind('<Enter>', lambda event: cancel_button_hover_in(event, top_canvas))
cancel_button.bind('<Leave>', lambda event: cancel_button_hover_out(event, top_canvas))

# Get Time Now button
time_now_icon = ImageTk.PhotoImage(Image.open('images/t_now.png').resize((45, 45), Image.ANTIALIAS))
time_now_button = Button(top_canvas, image=time_now_icon, command=lambda: get_current_date_time(top_canvas))
tn_window = top_canvas.create_window(t_time_now_button_x, t_time_now_button_y, window=time_now_button)
time_now_button.bind('<Enter>', lambda event: time_now_button_hover_in(event, top_canvas))
time_now_button.bind('<Leave>', lambda event: time_now_button_hover_out(event, top_canvas))



# MIDDLE CANVAS---------------------------------------------------------------------------------------

middle_canvas = Canvas(root, width=ROOT_WIDTH, height=200, borderwidth=0)
middle_canvas.grid(row=1, column=0, columnspan=2, sticky=W + E)
img2 = ImageTk.PhotoImage(Image.open('images/time_dark.jfif').resize((600, 400)), Image.ANTIALIAS)
middle_canvas.create_image(0, 0, image=img2, anchor=NW)
middle_canvas.create_text(m_text_x, m_text_y, text='Date/Time Completed',
                          font=('android 7', 25), fill='#e8ebea')

# Middle Canvas Hour Spin Box
m_hour_spin_box = Spinbox(middle_canvas, from_=0, to=23, width=2, font=('consolas', 20), fg='gray')
m_hour_spin_box.bind('<Enter>', lambda event: spin_box_hover_in(event, m_hour_spin_box))
m_hour_spin_box.bind('<Leave>', lambda event: spin_box_hover_out(event, m_hour_spin_box))
middle_canvas.create_window(290, 100, window=m_hour_spin_box)

# Colon separating minute from hours
middle_canvas.create_text(323, 100, text=':', font=('android 7', 25, 'bold'), fill='#e8ebea')

# Middle Canvas Minute Spin Box
m_minute_spin_box = Spinbox(middle_canvas, from_=0, to=59, width=2, font=('consolas', 20), fg='gray')
middle_canvas.create_window(355, 100, window=m_minute_spin_box)
m_minute_spin_box.bind('<Enter>', lambda event: spin_box_hover_in(event, m_minute_spin_box))
m_minute_spin_box.bind('<Leave>', lambda event: spin_box_hover_out(event, m_minute_spin_box))

# Time text
middle_canvas.create_text(m_time_text_x, m_time_text_y, text='Time:', font=('consolas', 20), fill='#e8ebea')

# Date text
middle_canvas.create_text(m_date_text_x, m_date_text_y, text='Date:', font=('consolas', 20), fill='#e8ebea')

# Calendar button for middle canvas
m_cal_icon = ImageTk.PhotoImage(Image.open('images/cal2_icon.png').resize((45, 45), Image.ANTIALIAS))
m_calendar_button = Button(middle_canvas, image=m_cal_icon, bg='light gray',
                           command=lambda: open_calendar(middle_canvas))
cl_window2 = middle_canvas.create_window(m_cal_button_x, m_cal_button_y, window=m_calendar_button)
m_calendar_button.bind('<Enter>', lambda event: calendar_button_hover_in(event, middle_canvas))
m_calendar_button.bind('<Leave>', lambda event: calendar_button_hover_out(event, middle_canvas))

# Selected date and time text
s_time2 = ''
selected_date_label2 = Label(middle_canvas, text='You selected: ' + date_from_calendar2 + ' ' + s_time2,
                             font=('consolas', 20, 'italic', 'bold'))

# Checkmark icon
c_icon2 = ImageTk.PhotoImage(Image.open('images/check_icon.png').resize((60, 60), Image.ANTIALIAS))
check_button2 = Button(middle_canvas, image=c_icon2, command=lambda: confirm_button_click(middle_canvas))
cf_window2 = middle_canvas.create_window(m_cf_button_x, m_cf_button_y, window=check_button2)
check_button2.bind('<Enter>', lambda event: confirm_button_hover_in(event, middle_canvas))
check_button2.bind('<Leave>', lambda event: confirm_button_hover_out(event, middle_canvas))

# Cancel icon
cancel_icon2 = ImageTk.PhotoImage(Image.open('images/remove_icon.jpg').resize((60, 60), Image.ANTIALIAS))
cancel_button2 = Button(middle_canvas, image=cancel_icon2, command=lambda: cancel_button_click(middle_canvas))
cx_window2 = middle_canvas.create_window(m_cx_button_x, m_cx_button_y, window=cancel_button2)
cancel_button2.bind('<Enter>', lambda event: cancel_button_hover_in(event, middle_canvas))
cancel_button2.bind('<Leave>', lambda event: cancel_button_hover_out(event, middle_canvas))

# Get Time Now button
time_now_icon2 = ImageTk.PhotoImage(Image.open('images/t_now.png').resize((45, 45), Image.ANTIALIAS))
time_now_button2 = Button(middle_canvas, image=time_now_icon2, command=lambda: get_current_date_time(middle_canvas))
tn_window2 = middle_canvas.create_window(m_time_now_button_x, m_time_now_button_y, window=time_now_button2)
time_now_button2.bind('<Enter>', lambda event: time_now_button_hover_in(event, middle_canvas))
time_now_button2.bind('<Leave>', lambda event: time_now_button_hover_out(event, middle_canvas))



# BOTTOM CANVAS-----------------------------------------------------------------------------------

bottom_canvas = Canvas(root, width=ROOT_WIDTH, height=200, borderwidth=0)
bottom_canvas.grid(row=2, column=0, columnspan=2, sticky=W + E)
img3 = ImageTk.PhotoImage(Image.open('images/money_dark.jfif').resize((600, 400)), Image.ANTIALIAS)
bottom_canvas.create_image(0, 0, image=img3, anchor=NW)
bottom_canvas.create_text(bottom_text_x, bottom_text_y, text='Payment', font=('android 7', 25), fill='#e8ebea')

# Calculate Payment Button
calc_payment_button = Button(bottom_canvas, text='Calculate Payment', font=('consolas', 20),
                             command=payment_calculation)
bottom_canvas.create_window(cbutton_x, cbutton_y, window=calc_payment_button)
calc_payment_button.bind('<Enter>', calculate_payment_button_hover_in)
calc_payment_button.bind('<Leave>', calculate_payment_button_hover_out)

# Reset Button
reset_icon = ImageTk.PhotoImage(Image.open('images/Reset.png').resize((60, 60), Image.ANTIALIAS))
reset_button = Button(bottom_canvas, image=reset_icon, command=reset_values)
r_window = bottom_canvas.create_window(rbutton_x, rbutton_y, window=reset_button)
reset_button.bind('<Enter>', reset_button_hover_in)
reset_button.bind('<Leave>', reset_button_hover_out)

# Save to Excel Button
save_icon = ImageTk.PhotoImage(Image.open('images/save.png').resize((65, 60), Image.ANTIALIAS))
save_button = Button(bottom_canvas, image=save_icon, command=save_to_excel)
sb_window = bottom_canvas.create_window(sbutton_x, sbutton_y, window=save_button)
save_button.bind('<Enter>', save_button_hover_in)
save_button.bind('<Leave>', save_button_hover_out)



# STATUS BAR--------------------------------------------------------------------------------
global left_status_text, date_and_time_text, status_bar

# Left Status bar text
status_bar_left = Label(root, text=rate_text, font=('consolas', 12), bg='#5fc29e', anchor=W)
status_bar_left.grid(row=3, column=0, sticky=W)

# Right status bar text
status_bar_right = Label(root, text='', font=('consolas', 12), bg='#5fc29e', anchor=E)
status_bar_right.grid(row=3, column=1, sticky=E)

# Thread for status bar time
threading.Thread(target=status_bar_time_update).start()



# MAIN--------------------
if __name__ == '__main__':
    root.mainloop()
    # TODO: 1. Use global coords for spin boxes
    #       2. Use global coords for colons
